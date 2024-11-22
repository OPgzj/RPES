from PySide6.QtWidgets import QMainWindow, QProgressDialog, QMessageBox, QSplitter, QTableWidgetItem, QScrollArea, QTableWidget, QLineEdit, QGraphicsPixmapItem, QFileDialog, QPushButton, QGraphicsView, QGraphicsScene, QHBoxLayout, QToolBox, QLabel, QVBoxLayout, QWidget
from PySide6.QtGui import QPixmap, QImage
from PySide6.QtCore import Qt, QTimer
import os, sys, cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
import shutil

from detect_dual import Detector


# 生成资源文件目录访问路径
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):  # 是否Bundle Resource
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class CrackDetectWindow(QMainWindow): 
    
    def __init__(self):
        super().__init__()
        # loadUi('ui/main_windowUI.ui', self)
    
        self.originalImagePath = ''

        # self.default_model_path = 'E:\\ricePhenoExt\\RPES\\yologui\\My_Work_Place\\weights\\best5_1.pt'
        self.default_model_path = resource_path(os.path.join("res", "best5_1.pt"))
        self.detector = Detector(self.default_model_path)
        self.progressDialog = None  
        self.showProgressDialog = False  # 进度弹窗显示控制
        # self.detectionThread = None
        
        self.setupUi()
        self.initUI()
    
    def load_stylesheet(self):
        # with open("E:/ricePhenoExt/RPES/yologui/resource/style.qss", "r", encoding='utf-8') as file:
        with open(resource_path(os.path.join("res", "style.qss")), "r", encoding='utf-8') as file:
            self.setStyleSheet(file.read())
    
    # 一、布局
    def setupUi(self):
        # 1. 设置界面：分为page1（单图）和page2（批量）
        self.load_stylesheet()
        self.setWindowTitle('High-throughput Soybean Seed Cracks Detecting System')
        self.setGeometry(50, 50, 900, 800)
        
        toolBox = QToolBox()

        self.page1 = QWidget()
        page1Layout = QVBoxLayout()
        self.page2 = QWidget()
        page2Layout = QVBoxLayout()
        
        self.page1.setObjectName("page1")
        self.page2.setObjectName("page2")


        # 2. Page1 - 单图识别
        # 上半部分布局
        topLayout = QHBoxLayout()
        # 左-图片显示区域
        self.imageLabel = QLabel('')
        self.imageLabel.setFixedSize(640, 640)  # 调整大小以适应窗口
        topLayout.addWidget(self.imageLabel)
        # 图形视图设置
        self.scene = QGraphicsScene()
        self.view = QGraphicsView(self.scene)
        self.imageLabel.setLayout(QVBoxLayout())
        self.imageLabel.layout().addWidget(self.view)
        
        # 右-信息显示区域
        # 创建显示检测信息的标签
        self.totalBeansLabel = QLabel('Total Beans:')
        self.crackedBeansLabel = QLabel('Crack Number:')
        self.crackRatioLabel = QLabel('Crack Ratio:')
        self.detectionTimeLabel = QLabel('Detection Time:')
        # 创建一个垂直布局用于放置上述标签
        infoLayout = QVBoxLayout()
        infoLayout.addWidget(self.totalBeansLabel)
        infoLayout.addWidget(self.crackedBeansLabel)
        infoLayout.addWidget(self.crackRatioLabel)
        infoLayout.addWidget(self.detectionTimeLabel)
        # 创建一个容器（Widget）并设置布局
        infoWidget = QWidget()
        infoWidget.setLayout(infoLayout)
        # infoWidget.setFixedWidth(225)  # 设置固定宽度
        # 将容器（Widget）添加到topLayout
        topLayout.addWidget(infoWidget)

        # 下半部分布局 - 按钮
        bottomLayout = QHBoxLayout()
        self.uploadButton = QPushButton('Upload Image')
        self.detectButton = QPushButton('Detect')
        self.originalButton = QPushButton('Show Original Image')
        self.clearButton = QPushButton('Clear')

        bottomLayout.addWidget(self.uploadButton)
        bottomLayout.addWidget(self.detectButton)
        bottomLayout.addWidget(self.originalButton)
        bottomLayout.addWidget(self.clearButton)

        # 3. Page2 - 批量检测
        # 页面2 - 批量检测
        # (1) 创建组件
        # QSplitter
        self.splitter = QSplitter(Qt.Orientation.Vertical)
        # 按钮
        self.page2SelectFolderBtn = QPushButton('Select Folder to Detect')
        self.page2FolderPathLineEdit = QLineEdit('yologui/My_Work_Place/test_img/30imgtest')
        self.page2SelectSavePathBtn = QPushButton('Select Result Save Path')
        self.page2SavePathLineEdit = QLineEdit('yologui/My_Work_Place/test_results')
        self.page2StartDetectBtn = QPushButton('Start Detection')
        self.saveExcelButton = QPushButton("Save to Excel", self)
        self.page2openSavePathButton = QPushButton("View Save Path", self)
        self.page2clearButton = QPushButton('Clear Interface')
        self.page2clearFileButton = QPushButton('Clear Result Folder',self)
        # 表格
        self.resultsTable = QTableWidget()
        self.resultsTable.setColumnCount(5)  # 列数
        self.resultsTable.setHorizontalHeaderLabels(['Image Name', 'Total Beans', 'Crack Number', 'Crack Ratio', 'Time Used'])
        # 统计信息标签
        self.totalBeansLabel2 = QLabel("Total Beans in All Images:")
        self.totalCrackedBeansLabel2 = QLabel("Total Cracks in All Images:")
        self.averageCrackRatioLabel2 = QLabel("Average Crack Ratio:")
        self.averageDetectionTimeLabel2 = QLabel("Average Time Used:")
        self.totalBeansValueLabel2 = QLabel("0")
        self.totalCrackedBeansValueLabel2 = QLabel("0")
        self.averageCrackRatioValueLabel2 = QLabel("0%")
        self.averageDetectionTimeValueLabel2 = QLabel("0ms")
        # 滚动区域
        self.scrollArea = QScrollArea()
        self.scrollArea.setWidgetResizable(True)  # 允许滚动区域内的控件根据需要调整大小
        self.scrollArea.setMinimumHeight(400)  # 滚动区域的高度
        
        # （2）布局
        # 按钮部分
        buttonWidget = QWidget()  # 包含按钮和路径输入框
        buttonLayout = QVBoxLayout()  # 垂直布局
        # 文件夹选择
        page2FolderLayout = QHBoxLayout()
        page2FolderLayout.addWidget(self.page2SelectFolderBtn)
        page2FolderLayout.addWidget(self.page2FolderPathLineEdit)
        buttonLayout.addLayout(page2FolderLayout)
        # 保存路径选择
        page2SavePathLayout = QHBoxLayout()
        page2SavePathLayout.addWidget(self.page2SelectSavePathBtn)
        page2SavePathLayout.addWidget(self.page2SavePathLineEdit)
        buttonLayout.addLayout(page2SavePathLayout)
        # 检测、保存、清空、查看、清空本地按钮
        buttonLayout.addWidget(self.page2StartDetectBtn)
        buttonLayout.addWidget(self.saveExcelButton)  # 将按钮添加到布局中，确保使用您的布局变量名
        buttonLayout.addWidget(self.page2clearButton)
        buttonLayout.addWidget(self.page2openSavePathButton)
        buttonLayout.addWidget(self.page2clearFileButton)
        # 以上，放入容器
        buttonWidget.setLayout(buttonLayout)

        # 统计信息标签
        statsLayout = QVBoxLayout()
        statsLayout.addWidget(self.totalBeansLabel2)
        statsLayout.addWidget(self.totalBeansValueLabel2)
        statsLayout.addWidget(self.totalCrackedBeansLabel2)
        statsLayout.addWidget(self.totalCrackedBeansValueLabel2)
        statsLayout.addWidget(self.averageCrackRatioLabel2)
        statsLayout.addWidget(self.averageCrackRatioValueLabel2)
        statsLayout.addWidget(self.averageDetectionTimeLabel2)
        statsLayout.addWidget(self.averageDetectionTimeValueLabel2)
        # 容器，包含表格和统计信息
        containerWidget = QWidget()
        containerLayout = QHBoxLayout()
        containerLayout.addWidget(self.resultsTable)  # 添加表格到容器布局
        containerLayout.addLayout(statsLayout)  # 添加统计信息布局到容器布局
        containerWidget.setLayout(containerLayout)
        self.scrollArea.setWidget(containerWidget)  # 将容器设置为滚动区域的子控件
        # 分隔条部分
        self.splitter = QSplitter(Qt.Orientation.Vertical)  # 创建一个垂直方向的分隔条
        self.splitter.addWidget(buttonWidget)  # 将包含按钮的小部件添加到分隔条上方
        self.splitter.addWidget(self.scrollArea)  # 将滚动区域添加到分隔条下方
        
        
        # 创建page2的布局，将分隔条添加到page2布局
        page2Layout = QVBoxLayout()
        page2Layout.addWidget(self.splitter)
        self.page2.setLayout(page2Layout)

        # 4. 整合布局
        page1Layout.addLayout(topLayout)
        page1Layout.addLayout(bottomLayout)
        self.page1.setLayout(page1Layout)
        self.page2.setLayout(page2Layout)
        toolBox.addItem(self.page1, '1. Single Image Detection')
        toolBox.addItem(self.page2, '2. Batch Detection')

        self.setCentralWidget(toolBox)
    
    # 二、连接信号与槽
    def initUI(self):
        # page1
        self.uploadButton.clicked.connect(self.uploadImage)
        self.detectButton.clicked.connect(self.detectCracks)
        self.originalButton.clicked.connect(self.showOriginalImage)
        self.clearButton.clicked.connect(self.clearImage)
        
        # page2
        self.page2SelectFolderBtn.clicked.connect(self.selectFolder)
        self.page2SelectSavePathBtn.clicked.connect(self.selectSavePath)
        self.page2StartDetectBtn.clicked.connect(self.startBatchDetection)
        self.saveExcelButton.clicked.connect(self.saveResultsToExcel)
        self.page2openSavePathButton.clicked.connect(self.openSavePath)
        self.page2clearButton.clicked.connect(self.clearImage2)
        self.page2clearFileButton.clicked.connect(self.clearDetectionResults)

        
        # other
        self.detector.progressUpdated.connect(self.handleProgressUpdated) # 进度信号，连接到批量检测弹窗
        # self.detectionThread.resultsReady.connect(self.startBatchDetection) # 显示信息信号，连接到批量检测按钮
    
    # 三、槽函数-功能实现
    # 1. 单图检测
    # (1) 上传图片按钮
    def uploadImage(self):
        # 上传图片
        imagePath, _ = QFileDialog.getOpenFileName(self, "Please select an image", "yologui/My_Work_Place/test_img/1", "Image files (*.jpg *.jpeg *.png)")
        if imagePath:
            self.originalImagePath = imagePath  # 保存原图路径
            pixmap = QPixmap(imagePath)
            self.displayImage(pixmap)

    def displayImage(self, pixmap):
        # 显示图片
        self.scene.clear()
        self.pixmapItem = QGraphicsPixmapItem(pixmap)
        self.scene.addItem(self.pixmapItem)
        self.view.fitInView(self.pixmapItem, Qt.AspectRatioMode.KeepAspectRatio)
    
    # (2) 检测按钮
    def detectCracks(self):
        if self.originalImagePath:
            self.showProgressDialog = False  # 单图检测不显示进度弹窗
            # 检测，参数为图片
            images_dict, info_dict, time_dict = self.detector.Detect_Soybean_Crack(source=self.originalImagePath, nosave=True)
            # 显示图片
            first_image_key = next(iter(images_dict))
            first_image_data = images_dict[first_image_key]
            im0 = first_image_data.astype(np.uint8)
            im0_rgb = cv2.cvtColor(im0, cv2.COLOR_BGR2RGB)
            qim = QImage(im0_rgb.data, im0_rgb.shape[1], im0_rgb.shape[0], im0_rgb.strides[0], QImage.Format.Format_RGB888)
            pixmap = QPixmap.fromImage(qim)
            self.displayImage(pixmap)
            # 显示检测信息
            # 更新标签
            self.totalBeansLabel.setText(f"Total Beans: {info_dict[first_image_key]['豆子总数']}")
            self.crackedBeansLabel.setText(f"Crack Number: {info_dict[first_image_key]['含裂纹豆子数']}")
            self.crackRatioLabel.setText(f"Crack Ratio: {info_dict[first_image_key]['裂纹占比'] * 100:.0f}%")
            self.detectionTimeLabel.setText(f"Detection Time: {time_dict[first_image_key]}")
        else:
            # 弹出未选择图片警告
            QMessageBox.warning(self, "Warning", "Please upload an image first!")

    # (3) 显示原图按钮
    def showOriginalImage(self):
        # 检查是否有原图路径存储
        if self.originalImagePath:
            pixmap = QPixmap(self.originalImagePath)
            self.displayImage(pixmap)
        else:
            # 如果没有图片路径，提示用户上传图片
            QMessageBox.warning(self, "Warning", "Please upload an image first!")

    # (4) 清空图片按钮
    def clearImage(self):
        # 清空图片
        self.scene.clear()
        self.originalImagePath = ''
        self.totalBeansLabel.setText(f"Total Beans: ")
        self.crackedBeansLabel.setText(f"Crack Number: ")
        self.crackRatioLabel.setText(f"Crack Ratio:  %")
        self.detectionTimeLabel.setText(f"Detection Time:")
        
    # 2. 批量检测
    # (1) 选择文件夹
    def selectFolder(self):
        folderPath = QFileDialog.getExistingDirectory(self, "Select Folder to Detect", self.page2FolderPathLineEdit.text())
        if folderPath:
            self.page2FolderPathLineEdit.setText(folderPath)

    # (2) 选择路径
    def selectSavePath(self):
        savePath = QFileDialog.getExistingDirectory(self, "Select Save Path", self.page2SavePathLineEdit.text())
        if savePath:
            self.page2SavePathLineEdit.setText(savePath)

    # (3) 批量检测按钮
    def startBatchDetection(self):
        sourcePath = self.page2FolderPathLineEdit.text()
        savePath = self.page2SavePathLineEdit.text()
        
        # 检测
        self.showProgressDialog = True  # 批量检测需要显示进度弹窗
        if self.showProgressDialog:
            if self.progressDialog is None:
                self.progressDialog = QProgressDialog("Detection Progress", "Stop", 0, 100, self)
                self.progressDialog.setStyleSheet("")
                self.progressDialog.setWindowTitle("Detecting...")
                self.progressDialog.setModal(True)
                self.progressDialog.setAutoReset(True)  # 确保完成时重置
                self.progressDialog.setAutoClose(True)  # 确保完成时关闭
                self.progressDialog.setMinimumDuration(0)  # 立即显示，无延迟
                self.progressDialog.setValue(0)  # 初始化进度为 0%
                # 连接取消按钮的信号
                self.progressDialog.canceled.connect(self.onCancel)
            self.progressDialog.show()
        images_dict, info_dict, time_dict = self.detector.Detect_Soybean_Crack(source=sourcePath, project=savePath)
        self.finishDetection()
        
        # 清空表格中的旧数据
        self.resultsTable.setRowCount(0)
        # 遍历字典并填充表格
        for index, path in enumerate(info_dict.keys(), start=1):
            row_position = self.resultsTable.rowCount()
            self.resultsTable.insertRow(row_position)  # 在表格末尾插入新行
            # 获取当前图片的检测信息、检测时间、图片名
            info = info_dict[path]
            detection_time = time_dict[path]  # 从 time_dict 获取检测时间
            filename = os.path.basename(path)
            # 填充数据
            self.resultsTable.setItem(row_position, 0, QTableWidgetItem(filename))
            self.resultsTable.setItem(row_position, 1, QTableWidgetItem(str(info["豆子总数"])))
            self.resultsTable.setItem(row_position, 2, QTableWidgetItem(str(info["含裂纹豆子数"])))
            self.resultsTable.setItem(row_position, 3, QTableWidgetItem(f"{info['裂纹占比'] * 100:.2f}%"))
            self.resultsTable.setItem(row_position, 4, QTableWidgetItem(detection_time))  # 使用正确的检测时间
            
        # 总信息计算
        total_beans = sum(info["豆子总数"] for info in info_dict.values())
        total_cracked_beans = sum(info["含裂纹豆子数"] for info in info_dict.values())
        average_crack_ratio = sum(info["裂纹占比"] for info in info_dict.values()) / len(info_dict) if info_dict else 0
        average_detection_time = sum(float(time[:-2]) for time in time_dict.values()) / len(time_dict) if time_dict else 0

        # 更新统计信息标签
        self.totalBeansValueLabel2.setText(str(total_beans))
        self.totalCrackedBeansValueLabel2.setText(str(total_cracked_beans))
        self.averageCrackRatioValueLabel2.setText(f"{average_crack_ratio * 100:.2f}%")
        self.averageDetectionTimeValueLabel2.setText(f"{average_detection_time:.1f}ms")

    # 批量检测进度弹窗
    def handleProgressUpdated(self, progress):
        print("Progress updated to:", progress)  # 打印进度，查看是否被调用
        if self.progressDialog and self.showProgressDialog:
            # 更新UI进度条等
            self.progressDialog.setValue(progress)
    
    def finishDetection(self):
        if self.progressDialog and self.showProgressDialog:
            self.progressDialog.close()
        self.showProgressDialog = False  # 重置标志
        
    def onCancel(self):
        # 设置一个标志，以通知检测过程需要停止
        self.detector.stop_requested = True
        # 使用 QTimer 设置500毫秒后重置 stop_requested 标志
        QTimer.singleShot(500, lambda: setattr(self.detector, 'stop_requested', False))
    
    # 追加：保存按钮
    def saveResultsToExcel(self):
        # 读取保存路径
        savePath = self.page2SavePathLineEdit.text()
        excel_path = os.path.join(savePath, "Detection_Results.xlsx")
        
        # 创建Excel工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Results"
        # 总信息的DataFrame
        total_info = pd.DataFrame({
            "Total Beans in All Images": [self.totalBeansValueLabel2.text()],
            "Total Cracks in All Images": [self.totalCrackedBeansValueLabel2.text()],
            "Average Crack Ratio": [self.averageCrackRatioValueLabel2.text()],
            "Average Time Used": [self.averageDetectionTimeValueLabel2.text()]
        })

        # 写入总信息
        ws.append(["Detection Information Summary"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(bold=True)
        
        ws.append(total_info.columns.tolist())
        for r in dataframe_to_rows(total_info, index=False, header=False):
            ws.append(r)

        # 空行
        ws.append([])

        # 每张图的检测信息
        ws.append(["Detection Information for Each Image"])
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=5)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A5'].font = Font(bold=True)
        
        columns = ["Image Name", "Bean Count", "Cracked Bean Count", "Crack Ratio", "Detection Time"]
        ws.append(columns)

        # 从表格中读取数据
        for row in range(self.resultsTable.rowCount()):
            row_data = [
                self.resultsTable.item(row, i).text() for i in range(self.resultsTable.columnCount())
            ]
            ws.append(row_data)
        
        # 保存Excel文件
        wb.save(excel_path)
        
        # 弹出消息框通知用户保存已完成
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("Save Complete")
        msg_box.setText("Excel file has been successfully saved!")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    # 追加：打开保存路径
    def openSavePath(self):
        save_path = self.page2SavePathLineEdit.text()
        save_path = save_path.replace('/', '\\')
        if os.path.exists(save_path):  # 确保路径存在
            if os.name == 'nt':  # 对于Windows
                os.startfile(save_path)
            elif os.name == 'posix':  # 对于macOS, Linux
                subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', save_path])
        else:
            self.showErrorMessage("Path does not exist, please check!")
    
    def showErrorMessage(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle("error")
        msg_box.setText(message)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)
        msg_box.exec()
    
    # (4) 清空按钮
    def clearImage2(self):
        self.resultsTable.setRowCount(0)
        self.totalBeansValueLabel2.setText(f"")
        self.totalCrackedBeansValueLabel2.setText(f"")
        self.averageCrackRatioValueLabel2.setText(f"")
        self.averageDetectionTimeValueLabel2.setText(f"")
        
    # 追加：清空结果文件夹
    def clearDetectionResults(self):
        save_path = self.page2SavePathLineEdit.text()
        if os.path.exists(save_path):  # 确保路径存在
            # 确认操作
            if self.confirmClear():
                for filename in os.listdir(save_path):
                    file_path = os.path.join(save_path, filename)
                    try:
                        if os.path.isfile(file_path) or os.path.islink(file_path):
                            os.unlink(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                    except Exception as e:
                        self.showErrorMessage(f"Unable to delete {filename}：{str(e)}")
                        return
                self.showInfoMessage("All detection results have been successfully cleared!")
            else:
                self.showInfoMessage("Clear operation has been cancelled.")
        else:
            self.showErrorMessage("The specified path does not exist, please check!")

    def confirmClear(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Warning)
        msg_box.setWindowTitle("Confirm Clear")
        msg_box.setText("Are you sure you want to delete all files in the save path?")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setDefaultButton(QMessageBox.StandardButton.No)
        response = msg_box.exec()
        return response == QMessageBox.StandardButton.Yes
    
    def showInfoMessage(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("message")
        msg_box.setText(message)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)
        msg_box.exec()

